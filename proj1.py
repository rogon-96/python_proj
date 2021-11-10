import numpy as np
import pandas as pd
import os
import openpyxl

slist = pd.read_csv('./sample_input/master_roll.csv')
response = pd.read_csv('./sample_input/responses.csv')

# crtopts = response['Answer','Roll Number']
crt = [response.iat[0,j] for j in range(7,35)]

if not os.path.exists('./out/'):
	os.mkdir('./out/')

for i in range(len(response)):
    
    wb =  openpyxl.Workbook() #creation and initialising a workbook
    sheet = wb.active
    sheet.title = response.at[i,'Roll Number']
    sheet.merge_cells('A1:E4')
    # add image here
    sheet.merge_cells('A5:E5')
    sheet.cell(row=5,column=1).value = "Mark Sheet"
    sheet.cell(row=6,column=1).value = "Name:"
    sheet.merge_cells('B6:C6')
    sheet.cell(row=6,column=2).value = slist.at[i,'name']
    sheet.cell(row=6,column=4).value = "Exam:"
    sheet.cell(row=6,column=5).value = "quiz"
    sheet.cell(row=7,column=1).value = "Roll Number:"
    sheet.cell(row=7,column=2).value = response.at[i,'Roll Number']

    sheet.cell(row=9,column=2).value = "Right"
    sheet.cell(row=9,column=3).value = "Wrong"
    sheet.cell(row=9,column=4).value = "Not Attempt"
    sheet.cell(row=9,column=5).value = "Max"
    sheet.cell(row=10,column=1).value = "No."
    sheet.cell(row=11,column=1).value = "Marking"
    sheet.cell(row=12,column=1).value = "Total"

    r,c = 16,1
    rgt,mrks,wrg,wmrks,na = 0,5,0,-1,0
    sheet.cell(row=15,column=1).value = "Student Ans"
    sheet.cell(row=15,column=2).value = "Correct Ans"
    sheet.cell(row=15,column=4).value = "Student Ans"
    sheet.cell(row=15,column=5).value = "Correct Ans"

    lt = []
    for j in range(7,35) : #Checking the Options
        if(j==32): r,c=16,4
        sheet.cell(row=r,column=c).value = response.iat[i,j]
        sheet.cell(row=r,column=c+1).value = crt[j-7]
        if(response.iat[i,j]==crt[j-7]):#counting right,wrong and not attempted answers 
            rgt+=1
        else: wrg+=1
        r+=1
        lt.append(response.iat[i,j])
        # print(response.iat[i,j])

    na = pd.Series(lt).isna().sum()
    wrg-=na
    sheet.cell(row=10,column=2).value = rgt
    sheet.cell(row=11,column=2).value = mrks
    sheet.cell(row=12,column=2).value = rgt*mrks
    sheet.cell(row=10,column=3).value = wrg
    sheet.cell(row=11,column=3).value = wmrks
    sheet.cell(row=12,column=3).value = wrg*wmrks
    sheet.cell(row=10,column=4).value = na
    sheet.cell(row=11,column=4).value = 0
    sheet.cell(row=10,column=5).value = rgt+wrg+na
    sheet.cell(row=12,column=5).value = str(rgt*mrks+wrg*wmrks)+'/'+str(140)
    wb.save("./out/"+sheet.title+'.xlsx')
